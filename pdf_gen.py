## References ##
## https://pyfpdf.readthedocs.io

######## import statements ########
import pandas as pd
import numpy as np
from fpdf import FPDF
from openpyxl import Workbook
from fpdf import FPDF
from datetime import date
import os

######## define variables ########

# today's date
today = date.today()

# Auto-generate invoice cycle date
invoice_cycle = today.strftime("%B %d, %Y") # another format: str(today.month) + ' / ' + str(today.year)

# edge case 1 of 2: check if source data is empty
filesize = os.path.getsize('data.xlsx')
if filesize == 0:
  print("The file is empty: " + str(filesize))
else:
# read in excel file and save as a dataframe
  data = pd.read_excel('data.xlsx')

df = pd.DataFrame(data)

# edge case 2 of 2: replace empty cells with NaN
df = df.replace(r'^\s*$', np.NaN, regex=True)

######## PDF Generator Class ########

# Class to generate PDF document and handle formatting
class Gen_PDF(FPDF):
  def header(self):
    pdf.set_xy(0, 0)
    pdf.set_font('arial', 'B', 14)
    pdf.cell(60)
    # Create the invoice header (1st row)
    pdf.cell(70, 10, 'Invoice for: ' + invoice_cycle, 0, 2, 'C')
    pdf.cell(-50)
    pdf.cell(50, 10, 'Index Column', 1, 0, 'C')
    pdf.cell(40, 10, 'Invoice Number', 1, 0, 'C')
    pdf.cell(40, 10, 'Company', 1, 0, 'C')
    pdf.cell(40, 10, 'Invoice Amount', 1, 2, 'C')


######## Instantiate the Gen_PDF class ########
pdf = Gen_PDF()
pdf.alias_nb_pages()
pdf.add_page()
pdf.set_font('Times', '', 12)
# left-aligns each row with the header
pdf.set_xy(100, 20)

######## Loop through and load data ########
for i in range(0, len(df)-1):
  col_ind = str(i)
  col_a = str(df["Invoice Number"][i])
  col_b = str(df["Company"][i])
  col_c =  "${:,.2f}".format(df["Invoide Amount"][i])
  pdf.cell(-90)
  pdf.cell(50, 10, '%s' % (col_ind), 1, 0, 'C')
  pdf.cell(40, 10, '%s' % (col_a), 0, 0, 'C')
  pdf.cell(40, 10, '%s' % (col_b), 0, 0, 'C')
  pdf.cell(40, 10, '%s' % (col_c), 0, 2, 'C')
  pdf.cell(-40)

# saves the file as invoices_'today's date'.pdf
pdf.output('invoice_{}.pdf'.format(today), 'F')
